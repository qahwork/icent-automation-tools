import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# 날짜 파싱 함수 (엑셀 날짜 변환)
def parse_excel_date(date_str):
    if not date_str or date_str.strip() in ["-", "N"]:
        return "-"
    date_str = date_str.strip()
    formats = ["%d-%b-%y", "%d-%b-%Y", "%d-%B-%Y", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    print(f"⚠️ 날짜 변환 실패: {date_str}")
    return "-"

# ------------------------------------------------------------------------------
# 1. LineDetails CSV 파일(들) 처리
# 현재 디렉토리 내 "LineDetails"로 시작하고 ".csv"로 끝나는 모든 파일을 수집
linedetail_files = [f for f in os.listdir() if f.startswith("LineDetails") and f.endswith(".csv")]

csv_dataframes = []
file_dates = {}  # 각 파일에서 추출한 날짜를 저장 (디버깅용)

for filename in linedetail_files:
    try:
        with open(filename, "r", encoding="utf-8") as f:
            header_lines = [next(f) for _ in range(5)]
        # 헤더 5행 중 "Date"가 포함된 행을 검색하여 날짜를 추출
        # 예: "Date	9-Apr-25" 또는 "Date 9-Apr-25"
        for line in header_lines:
            if "Date" in line:
                parts = re.split(r'\t|\s{2,}', line.strip())
                if len(parts) >= 2:
                    file_dates[filename] = parts[1]
                break

        # 상단 5행을 건너뛰고 CSV 데이터 읽기
        df = pd.read_csv(filename, skiprows=5, dtype=str).fillna("")
        csv_dataframes.append(df)
    except Exception as e:
        print(f"[ERROR] '{filename}' 처리 중 오류 발생: {e}")

if not csv_dataframes:
    raise FileNotFoundError("LineDetails 관련 CSV 파일을 찾을 수 없습니다.")

# 여러 CSV 파일을 하나의 DataFrame으로 결합
linedetails_df = pd.concat(csv_dataframes, ignore_index=True)

# ------------------------------------------------------------------------------
# 2. Serial Map 생성 from LineDetails CSV 데이터
serial_map = {}
for _, row in linedetails_df.iterrows():
    serial = row.get("PAK/Serial Number", "").strip()
    status = row.get("Status", "").strip().upper()
    if not serial:
        continue

    # 기본값 설정: 각 필드가 비어있을 때만 이후 업데이트하도록 "-"로 초기화함
    entry = serial_map.setdefault(serial, {
        "LDoS": "-", 
        "서비스 종류": "-", 
        "ACTIVE 종료일": "-", 
        "SIGNED 종료일": "-", 
        "모델명PID": "-", 
        "계약번호\n(Contract)": "-"
    })

    # 추가: Description 열에서 "meraki" (대소문자 무시) 포함 여부 확인
    description = row.get("Description", "").strip()
    if "meraki" in description.lower():
        # meraki가 포함된 경우에는 LDoS만 업데이트
        ldos_raw = row.get("Last Date of Support", "").strip()
        new_ldos = parse_excel_date(ldos_raw)
        # 현재 LDoS가 비어있거나 "-"인 경우에만 업데이트 (그리고 새 값이 "-"가 아닐 때)
        if (not entry["LDoS"] or entry["LDoS"] == "-") and new_ldos != "-":
            entry["LDoS"] = new_ldos
        # 다른 필드는 업데이트하지 않고 다음 row로 넘어감
        continue

    # 일반적으로 업데이트하는 경우: 각 필드는 새 값이 있을 때만 업데이트함

    # LDoS 업데이트: 새 값이 "-"가 아니거나 현재 값이 비어있을 경우만 업데이트
    ldos_raw = row.get("Last Date of Support", "").strip()
    new_ldos = parse_excel_date(ldos_raw)
    if new_ldos != "-" or entry["LDoS"] in ["", "-"]:
        entry["LDoS"] = new_ldos

    # 서비스 종류 업데이트: 값이 존재할 경우 업데이트
    offer_type = row.get("Service Level/Offer Type", "").strip()
    if offer_type:
        entry["서비스 종류"] = offer_type

    # End Date 업데이트 (ACTIVE 또는 SIGNED 상태에 따라)
    end_date_raw = row.get("End Date", "").strip()
    end_date_parsed = parse_excel_date(end_date_raw)
    if status == "ACTIVE":
        if end_date_parsed != "-" or entry["ACTIVE 종료일"] in ["", "-"]:
            entry["ACTIVE 종료일"] = end_date_parsed
    elif status == "SIGNED":
        if end_date_parsed != "-" or entry["SIGNED 종료일"] in ["", "-"]:
            entry["SIGNED 종료일"] = end_date_parsed

    # 모델명PID 업데이트
    model_name = row.get("Product /Offer Name", "").strip()
    if model_name:
        entry["모델명PID"] = model_name

    # 계약번호 업데이트
    contract_num = row.get("Subscription ID/Contract Number", "").strip()
    if contract_num:
        entry["계약번호\n(Contract)"] = contract_num
# ------------------------------------------------------------------------------
# 3. serials.csv와 LineDetails CSV의 시리얼 비교 (problem_serial.csv 생성)
serials_csv_path = os.path.join(os.getcwd(), "serials.csv")
if not os.path.exists(serials_csv_path):
    raise FileNotFoundError(f"serials.csv 파일을 찾을 수 없습니다. ({serials_csv_path})")
serials_df = pd.read_csv(serials_csv_path, dtype=str).fillna("")
# serials.csv는 'Serials' 컬럼에 시리얼 번호가 들어있다고 가정
serials_list = serials_df["Serial"].astype(str).str.strip().tolist()

# LineDetails에서 확인된 시리얼 집합
linedetail_serials = set(serial_map.keys())

# serials.csv에 있으나 LineDetails에 없는 시리얼 추출
problem_serials = [s for s in serials_list if s not in linedetail_serials]

# Filename 추가
serials_df['Filename'] = serials_df['Filename'].fillna("")  # Filename 컬럼이 없으면 빈 문자열 처리
problem_filenames = serials_df.loc[serials_df['Serial'].isin(problem_serials), 'Filename'].tolist()

# 문제 시리얼과 파일명 결합
problem_data = list(zip(problem_serials, problem_filenames))

# 문제 시리얼과 파일명을 포함한 DataFrame 생성
problem_df = pd.DataFrame(problem_data, columns=["Problem Serials", "Filename"])

# 문제 시리얼 CSV 파일 경로 지정
problem_csv_path = os.path.join(os.getcwd(), "problem_serial.csv")
problem_df.to_csv(problem_csv_path, index=False, encoding="utf-8-sig")
print(f"✅ 문제 시리얼 저장: {problem_csv_path}")

# ------------------------------------------------------------------------------
# 4. XLSX 파일 처리 및 업데이트 후 파일명 변경
# 엑셀 파일에서 갱신 대상 시트: "유지보수 대상장비"
def normalize(text):
    return str(text).replace("\n", "").replace(" ", "").strip() if text is not None else ""

# 각 컬럼 후보 지정
column_candidates = {
    "serial_col": ["시리얼"],
    "ldos_date_col": ["H/WLDoSDate", "H/WEOS(Support)날짜"],
    "service_type_col": ["서비스종류Subscription/ServiceLevel", "서비스종류"],
    "end_date_active_col": ["서비스종료일(Active)"],
    "end_date_signed_col": ["서비스종료일(SIGNED)"],
    "model_pid_col": ["모델명PID", "모델명\nPID"],
    "confirm_col": ["확인요청"],
    "contract_col": ["계약번호\n(Contract)", "Subscription ID/Contract Number"]
}

# 파일명 변경 함수: 파일명의 마지막 한글 문자 뒤에 _yyyy_mm_dd_업데이트 삽입
def rename_file_with_date(original_file):
    base, ext = os.path.splitext(original_file)
    # 마지막 한글 문자의 위치 찾기 (정규표현식 사용)
    matches = list(re.finditer(r"[가-힣]", base))
    # 현재 날짜 문자열 생성 (예: 2025_04_10)
    date_str = datetime.now().strftime("%Y_%m_%d")
    if matches:
        last_match = matches[-1]
        new_base = base[:last_match.end()] + f"_{date_str}_update"
    else:
        new_base = base + f"_{date_str}_update"
    new_name = new_base + ext
    os.rename(original_file, new_name)
    return new_name

# XLSX 파일 처리
for filename in os.listdir():
    if not filename.endswith(".xlsx"):
        continue

    try:
        wb = load_workbook(filename)
        if "유지보수 대상장비" not in wb.sheetnames:
            print(f"[SKIP] 시트 없음: {filename}")
            continue

        ws = wb["유지보수 대상장비"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {normalize(h): i for i, h in enumerate(headers)}

        col_indices = {}
        for key, candidates in column_candidates.items():
            found = None
            for c in candidates:
                norm_c = normalize(c)
                if norm_c in header_map:
                    found = header_map[norm_c]
                    break
            col_indices[key] = found

        # 헤더 매핑에서 확인요청(confirm_col)나 계약번호(contract_col)이 없으면,
        # 엑셀 파일에 새 열을 추가 (맨 마지막 열)
        max_index = len(headers)
        if col_indices["confirm_col"] is None:
            col_indices["confirm_col"] = max_index
            headers.append("확인요청")
            max_index += 1
        if col_indices["contract_col"] is None:
            col_indices["contract_col"] = max_index
            headers.append("계약번호\n(Contract)")
            max_index += 1

        # 필수 컬럼 헤더 매핑 체크 (serial_col, ldos_date_col, service_type_col, end_date_active_col, end_date_signed_col, model_pid_col)
        required_keys = ["serial_col", "ldos_date_col", "service_type_col", 
                         "end_date_active_col", "end_date_signed_col", "model_pid_col"]
        failed = False
        for k in required_keys:
            if col_indices[k] is None:
                print(f"\n[❌ 헤더 매핑 실패] {filename} - '{k}'에 해당하는 열을 찾을 수 없음 (후보: {column_candidates[k]})")
                failed = True
        if failed:
            print("엑셀 헤더 목록:")
            for i, h in enumerate(headers):
                print(f"  - [{i}] '{h}' → 정규화: '{normalize(h)}'")
            continue

        updated = 0
        debug_printed = False

        # 2행부터 데이터 처리
        for row in ws.iter_rows(min_row=2):
            serial_cell = row[col_indices["serial_col"]]
            serial = str(serial_cell.value).strip() if serial_cell.value is not None else ""
            if not serial:
                continue

            if serial in serial_map:
                info = serial_map[serial]
                if not debug_printed:
                    print(f"\n🔍 디버그 (1회) - {serial} → {info}")
                    debug_printed = True

                row[col_indices["ldos_date_col"]].value = info["LDoS"] if info["LDoS"] != "-" else "-"
                row[col_indices["service_type_col"]].value = info["서비스 종류"] if info["서비스 종류"] else "-"
                row[col_indices["end_date_active_col"]].value = info["ACTIVE 종료일"] if info["ACTIVE 종료일"] != "-" else "-"
                row[col_indices["end_date_signed_col"]].value = info["SIGNED 종료일"] if info["SIGNED 종료일"] != "-" else "-"
                row[col_indices["model_pid_col"]].value = info["모델명PID"] if info["모델명PID"] else "-"
                row[col_indices["contract_col"]].value = info["계약번호\n(Contract)"] if info["계약번호\n(Contract)"] else "-"
            else:
                # CSV에 해당 시리얼이 없으면 '확인요청' 열에 "CCW 검색불가" 기록
                row[col_indices["confirm_col"]].value = "CCW 검색불가"
            updated += 1

        wb.save(filename)
        print(f"✅ [UPDATED] {filename} - 총 {updated}건 업데이트 완료")

        # 파일명 변경: 파일명의 마지막 한글 뒤에 _yyyy_mm_dd_업데이트 추가
        new_filename = rename_file_with_date(filename)
        print(f"✅ 파일명 변경: {filename} → {new_filename}")

    except Exception as e:
        print(f"[ERROR] {filename} 처리 중 오류 발생: {e}")
