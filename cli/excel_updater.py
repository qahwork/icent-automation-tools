import os
import re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import logging
from configparser import ConfigParser

from core.utils import parse_excel_date, rename_file_with_date, normalize, get_logger

# 로거 설정
logger = get_logger("excel_updater")

# 설정 파일 읽기
config = ConfigParser()
config.read(os.path.join(os.path.dirname(__file__), '..', 'config', 'config.ini'))

# 설정값
header_skip_lines = config.getint('DEFAULT', 'header_skip_lines', fallback=5)
sheet_name = config.get('EXCEL', 'sheet_name', fallback="유지보수 대상장비")
output_dir = config.get('OUTPUT', 'output_dir', fallback="outputs")

# serial_map을 생성하는 함수 예시
def build_serial_map(linedetails_df):
    serial_map = {}
    for _, row in linedetails_df.iterrows():
        serial = row.get("PAK/Serial Number", "").strip()
        status = row.get("Status", "").strip().upper()
        if not serial:
            continue
        entry = serial_map.setdefault(serial, {
            "LDoS": "-",
            "서비스 종류": "-",
            "ACTIVE 종료일": "-",
            "SIGNED 종료일": "-",
            "모델명PID": "-",
            "계약번호": "-"
        })
        description = row.get("Description", "").strip()
        # meraki 체크 (대소문자 무시)
        if "meraki" in description.lower():
            ldos_raw = row.get("Last Date of Support", "").strip()
            new_ldos = parse_excel_date(ldos_raw)
            if (not entry["LDoS"] or entry["LDoS"] == "-") and new_ldos != "-":
                entry["LDoS"] = new_ldos
        model_name = row.get("Product /Offer Name", "").strip()
        if model_name:
            entry["모델명PID"] = model_name
            continue

        ldos_raw = row.get("Last Date of Support", "").strip()
        new_ldos = parse_excel_date(ldos_raw)
        if new_ldos != "-" or entry["LDoS"] in ["", "-"]:
            entry["LDoS"] = new_ldos

        offer_type = row.get("Service Level/Offer Type", "").strip()
        if offer_type:
            entry["서비스 종류"] = offer_type

        end_date_raw = row.get("End Date", "").strip()
        end_date_parsed = parse_excel_date(end_date_raw)
        if status == "ACTIVE":
            if end_date_parsed != "-" or entry["ACTIVE 종료일"] in ["", "-"]:
                entry["ACTIVE 종료일"] = end_date_parsed
        elif status == "SIGNED":
            if end_date_parsed != "-" or entry["SIGNED 종료일"] in ["", "-"]:
                entry["SIGNED 종료일"] = end_date_parsed

        model_name = row.get("Product /Offer Name", "").strip()
        if model_name:
            entry["모델명PID"] = model_name

        contract_num = row.get("Subscription ID/Contract Number", "").strip()
        if contract_num and contract_num.isdigit():
            entry["계약번호"] = int(contract_num)
        elif contract_num:
            logger.warning(f"⚠️ 계약번호가 숫자가 아니어서 변환되지 않았습니다: {contract_num}")
    return serial_map

def update_excel_files(serial_map):
    # 현재 폴더 내의 모든 Excel 파일 처리
    for filename in os.listdir():
        if not filename.endswith(".xlsx"):
            continue
        try:
            wb = load_workbook(filename)
            if sheet_name not in wb.sheetnames:
                logger.info(f"[SKIP] 시트 없음: {filename}")
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            # header_map: normalize 처리 후 인덱스를 매핑
            header_map = {normalize(h): i for i, h in enumerate(headers)}

            # 컬럼 후보는 config 파일이나 아래와 같이 직접 지정
            column_candidates = {
                "serial": config.get('EXCEL', 'serial_column', fallback="시리얼"),
                "ldos_date": config.get('EXCEL', 'ldos_date_candidates', fallback="").split(','),
                "service_type": config.get('EXCEL', 'service_type_candidates', fallback="").split(','),
                "end_date_active": config.get('EXCEL', 'end_date_active', fallback="서비스종료일(Active)"),
                "end_date_signed": config.get('EXCEL', 'end_date_signed', fallback="서비스종료일(SIGNED)"),
                "model_pid": config.get('EXCEL', 'model_pid', fallback="모델명PID"),
                "confirm": "확인요청",
                "contract": config.get('EXCEL', 'contract_column', fallback="계약번호")
            }

            # 컬럼 인덱스를 찾기
            col_indices = {}
            for key, cand in column_candidates.items():
                if isinstance(cand, list):
                    found = None
                    for c in cand:
                        norm_c = normalize(c)
                        if norm_c in header_map:
                            found = header_map[norm_c]
                            break
                    col_indices[key] = found
                else:
                    col_indices[key] = header_map.get(normalize(cand))
            # 업데이트 시작
            debug_printed = False
            for row in ws.iter_rows(min_row=2):
                serial_cell = row[col_indices["serial"]]
                serial = str(serial_cell.value).strip() if serial_cell.value is not None else ""
                if not serial:
                    continue
                if serial in serial_map:
                    info = serial_map[serial]
                    if not debug_printed:
                        logger.debug(f"Debug: {serial} → {info}")
                        debug_printed = True
                    # 업데이트 각 셀
                    for key in ["ldos_date", "service_type", "end_date_active", "end_date_signed", "model_pid"]:
                        if col_indices.get(key) is not None:
                            old_value = row[col_indices[key]].value
                            new_value = info.get({
                                "ldos_date": "LDoS",
                                "service_type": "서비스 종류",
                                "end_date_active": "ACTIVE 종료일",
                                "end_date_signed": "SIGNED 종료일",
                                "model_pid": "모델명PID"
                            }[key])
                            # 새 값이 "-" 또는 빈 값인 경우 기존 값 유지
                            if new_value != "-" and new_value and new_value != old_value:
                                row[col_indices[key]].value = new_value
                    # 계약번호 처리 (정수형 변환이 필요한 경우)
                    if col_indices.get("contract") is not None:
                        contract_info = info.get("계약번호")
                        if isinstance(contract_info, int):
                            row[col_indices["contract"]].value = contract_info
                        elif contract_info:
                            row[col_indices["contract"]].value = contract_info
                else:
                    # 없는 시리얼의 경우 '확인요청' 열에 기록
                    if col_indices.get("confirm") is not None:
                        row[col_indices["confirm"]].value = "CCW 검색불가"
            wb.save(filename)
            new_name = rename_file_with_date(filename)
            logger.info(f"✅ 파일 저장 완료: {new_name}")
        except Exception as e:
            logger.error(f"[ERROR] '{filename}' 처리 중 오류 발생: {e}")

def main():
    # 여기는 CSV 또는 다른 데이터 소스로부터 serial_map을 만드는 코드 예시입니다.
    # 실제 구현에서는 적절한 파일 읽기/병합 로직을 추가합니다.
    csv_files = [f for f in os.listdir() if f.startswith("LineDetails") and f.endswith(".csv")]
    if not csv_files:
        logger.error("LineDetails CSV 파일을 찾을 수 없습니다.")
        return

    csv_dataframes = []
    for filename in csv_files:
        try:
            # 헤더 5행 스킵 (config 기반으로 조정 가능)
            df = pd.read_csv(filename, skiprows=header_skip_lines, dtype=str).fillna("")
            csv_dataframes.append(df)
        except Exception as ex:
            logger.error(f"[ERROR] '{filename}' 처리 중 오류: {ex}")

    if not csv_dataframes:
        logger.error("CSV 데이터 읽기 실패.")
        return

    combined_df = pd.concat(csv_dataframes, ignore_index=True)
    serial_map = build_serial_map(combined_df)
    update_excel_files(serial_map)

if __name__ == '__main__':
    main()
